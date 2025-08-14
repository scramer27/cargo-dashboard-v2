import pandas as pd
import sys
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Border, Side
from log_parser import process_log_file
import numpy as np
import io

def create_plots(daily_data, day_str):
    """generates and saves distribution and throughput plots for a given day."""
    plot_paths = {}
    
    # --- time distribution plot ---
    stow_df = daily_data["stow_events"]
    retrieve_df = daily_data["retrieve_events"]
    read_label_df = daily_data["read_label_events"]

    stow_times = stow_df["Time (s)"].dropna() if not stow_df.empty else pd.Series(dtype='float64')
    retrieve_times = retrieve_df["Time (s)"].dropna() if not retrieve_df.empty else pd.Series(dtype='float64')
    read_label_times = read_label_df["Time (s)"].dropna() if not read_label_df.empty else pd.Series(dtype='float64')
    
    fig1, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(20, 6), sharey=True)
    fig1.suptitle(f'Event Time Distribution for {day_str} ({daily_data["date"]})', fontsize=16)
    
    ax1.hist(stow_times, bins=10, color='blue', edgecolor='black')
    ax1.set_title('Stow Times')
    ax1.set_xlabel('Time (s)')
    ax1.set_ylabel('Frequency')

    ax2.hist(retrieve_times, bins=10, color='pink', edgecolor='black')
    ax2.set_title('Retrieval Times')
    ax2.set_xlabel('Time (s)')

    ax3.hist(read_label_times, bins=10, color='purple', edgecolor='black')
    ax3.set_title('Read Label Times')
    ax3.set_xlabel('Time (s)')
    
    dist_plot_path = f"{day_str}_distribution.png"
    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    plt.savefig(dist_plot_path)
    plt.close(fig1)
    plot_paths['distribution'] = dist_plot_path
    
    return plot_paths

def get_stats_df(df, column_name):
    """calculates statistics for a dataframe column and returns a new dataframe."""
    if df.empty or column_name not in df.columns or df[column_name].isnull().all():
        stats = {'Average': 'N/A', 'Minimum': 'N/A', 'Maximum': 'N/A'}
        return pd.DataFrame.from_dict(stats, orient='index', columns=['Time (s)'])

    times = pd.to_numeric(df[column_name], errors='coerce').dropna()
    
    if times.empty:
        stats = {'Average': 'N/A', 'Minimum': 'N/A', 'Maximum': 'N/A'}
    else:
        stats = {
            'Average': f"{np.mean(times):.2f}",
            'Minimum': f"{np.min(times):.2f}",
            'Maximum': f"{np.max(times):.2f}"
        }
    return pd.DataFrame.from_dict(stats, orient='index', columns=['Time (s)'])

def add_distribution_chart(writer, sheet_name, df, col_name, chart_title, start_row, start_col):
    """calculates histogram data, writes it to excel, and adds a chart."""
    if df.empty or df[col_name].isnull().all():
        return start_row

    times = pd.to_numeric(df[col_name], errors='coerce').dropna()
    if times.empty:
        return start_row

    counts, bin_edges = np.histogram(times, bins=10)
    bin_labels = [f'{edge:.2f}-{bin_edges[i+1]:.2f}' for i, edge in enumerate(bin_edges[:-1])]
    hist_df = pd.DataFrame({'Time Bins (s)': bin_labels, 'Frequency': counts})

    hist_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=start_col, index=False)
    
    # chart creation is handled in the final pass
    return start_row + len(hist_df) + 2

def apply_cell_style(cell):
    """applies a bold font and thin border to a cell."""
    if cell.value is None: return
    cell.font = Font(bold=True)
    cell.border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

def update_master_report(log_file_paths, excel_path, target_sheet_name):
    """processes logs and updates the master excel workbook with new data and charts."""
    file_exists = os.path.exists(excel_path)
    master_summary_df = pd.DataFrame()
    error_summary_df = pd.DataFrame()
    existing_day_sheets = {}

    if file_exists:
        try:
            xls = pd.ExcelFile(excel_path)
            if target_sheet_name in xls.sheet_names:
                master_summary_df = pd.read_excel(xls, sheet_name=target_sheet_name)
            if 'Error Summary' in xls.sheet_names:
                error_summary_df = pd.read_excel(xls, sheet_name='Error Summary')
            for sheet_name in xls.sheet_names:
                if sheet_name.startswith('Day '):
                    # read raw data; charts will be regenerated later
                    existing_day_sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        except Exception as e:
            print(f"Warning: Could not read existing report '{excel_path}'. A new one will be created. Error: {e}")
            master_summary_df = pd.DataFrame()
            error_summary_df = pd.DataFrame()
            existing_day_sheets = {}

    day_counter = (master_summary_df['Day'].max() + 1) if not master_summary_df.empty and 'Day' in master_summary_df.columns else 1
    
    new_daily_data = []
    for log_path in log_file_paths:
        print(f"Processing file: {log_path}...")
        daily_data = process_log_file(log_path)
        if not daily_data or not daily_data["date"]:
            print(f"Could not extract data or date from {log_path}. Skipping.")
            continue
        
        daily_data["summary"]["Day"] = day_counter
        if "error_summary" in daily_data and not daily_data["error_summary"].empty:
            daily_data["error_summary"]["Day"] = day_counter
        new_daily_data.append(daily_data)
        day_counter += 1

    if not new_daily_data:
        print("No new data to add to the report.")
        return

    all_summaries = [d['summary'] for d in new_daily_data]
    updated_summary_df = pd.concat([master_summary_df, pd.DataFrame(all_summaries)], ignore_index=True)
    
    all_error_summaries = [d['error_summary'] for d in new_daily_data if 'error_summary' in d]
    if all_error_summaries:
        updated_error_summary_df = pd.concat([error_summary_df] + all_error_summaries, ignore_index=True)
    else:
        updated_error_summary_df = error_summary_df

    for col in ['Stow Avg (s)', 'Retrieve Avg (s)', 'Read Label Avg (s)']:
        if col in updated_summary_df.columns:
            updated_summary_df[col] = updated_summary_df[col].apply(lambda x: f"{x:.2f}" if isinstance(x, (int, float)) else 'N/A')

    updated_summary_df.sort_values(by='Day', inplace=True)
    # define the column order for the master summary sheet
    cols = [
        'Day', 'Date', 
        'Pickup Avg (s)', 'Placement Avg (s)', 'Stow Avg (s)', 
        'Retrieve Avg (s)', 'Read Label Avg (s)', 
        'Packages Picked Up', 'Pickup Attempts',
        'Packages Placed', 'Placement Attempts', 
        'Packages Retrieved', 'Retrieval Attempts', 
        'Total Errors', 
        'Stow Driver Shift Time (hr)', 'Retrieve Driver Shift Time (hr)'
    ]
    updated_summary_df = updated_summary_df.reindex(columns=cols)

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Write the specific summary to its target sheet
        updated_summary_df.to_excel(writer, sheet_name=target_sheet_name, index=False)
        
        # Write other sheets that might exist
        book = writer.book
        for sheet_name, df in existing_day_sheets.items():
            if sheet_name not in book.sheetnames:
                 df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        if not updated_error_summary_df.empty:
            # reorder columns to have day and date first
            if 'Day' in updated_error_summary_df.columns:
                cols = ['Day', 'Date'] + [c for c in updated_error_summary_df.columns if c not in ['Day', 'Date']]
                updated_error_summary_df = updated_error_summary_df[cols]
            updated_error_summary_df.to_excel(writer, sheet_name='Error Summary', index=False)

        for daily_data in new_daily_data:
            sheet_name = f"Day {daily_data['summary']['Day']}"
            
            # Get all 4 event types
            pickup_df = daily_data.get("pickup_events", pd.DataFrame())
            placement_df = daily_data.get("placement_events", pd.DataFrame())
            retrieve_df = daily_data["retrieve_events"]
            read_label_df = daily_data["read_label_events"]
            failures_df = daily_data["failures"]

            # Write the main dataframes to the sheet with new layout
            pickup_df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, index=False)       # Column A-C: Pickup
            placement_df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=3, index=False)    # Column D-F: Placement
            retrieve_df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=6, index=False)     # Column G-I: Retrieve
            read_label_df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=9, index=False)  # Column J-L: Read Label
            
            max_len = max(len(pickup_df), len(placement_df), len(retrieve_df), len(read_label_df))
            current_row = max_len + 2
            
            # Calculate and write statistics for all 4 event types
            stats_pickup = get_stats_df(pickup_df, 'Time (s)')
            stats_placement = get_stats_df(placement_df, 'Time (s)')
            stats_retrieve = get_stats_df(retrieve_df, 'Time (s)')
            stats_read_label = get_stats_df(read_label_df, 'Time (s)')
            
            stats_pickup.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=0)
            stats_placement.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=3)
            stats_retrieve.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=6)
            stats_read_label.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=9)
            
            current_row += len(stats_pickup) + 2
            if not failures_df.empty:
                failures_df.to_excel(writer, sheet_name=sheet_name, startrow=current_row, index=False)
                current_row += len(failures_df) + 2
            
            # Write chart data for all 4 event types
            chart_start_row = current_row
            chart_start_row = add_distribution_chart(writer, sheet_name, pickup_df, 'Time (s)', 'Pickup Time Distribution', chart_start_row, 0)
            chart_start_row = add_distribution_chart(writer, sheet_name, placement_df, 'Time (s)', 'Placement Time Distribution', chart_start_row, 0)
            chart_start_row = add_distribution_chart(writer, sheet_name, retrieve_df, 'Time (s)', 'Retrieve Time Distribution', chart_start_row, 0)
            chart_start_row = add_distribution_chart(writer, sheet_name, read_label_df, 'Time (s)', 'Read Label Time Distribution', chart_start_row, 0)

    # --- final formatting and charting pass ---
    wb = load_workbook(excel_path)
    
    # style master summary
    if target_sheet_name in wb.sheetnames:
        ws_summary = wb[target_sheet_name]
        for col in range(1, len(updated_summary_df.columns) + 1):
            apply_cell_style(ws_summary.cell(row=1, column=col))

    # style and add charts to all day sheets
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('Day '):
            ws_day = wb[sheet_name]
            
            # main data headers: stow(4), retrieve(5), read label(3)
            for col in list(range(1, 5)) + list(range(6, 11)) + list(range(12, 15)):
                apply_cell_style(ws_day.cell(row=2, column=col))

            # find and style headers, and find chart data locations
            chart_data_rows = []
            for r in range(3, ws_day.max_row + 1):
                cell_val = ws_day.cell(row=r, column=1).value
                if cell_val == 'Average':
                    stats_row = r
                    for r_offset in range(3):
                        for c_base in [1, 6, 12]:
                            apply_cell_style(ws_day.cell(row=stats_row + r_offset, column=c_base))
                    for c_base in [2, 7, 13]:
                        apply_cell_style(ws_day.cell(row=stats_row - 1, column=c_base))
                elif cell_val == 'Timestamp' and ws_day.cell(row=r-1, column=1).value is None:
                    failures_row = r
                    for col in range(1, 4):
                        apply_cell_style(ws_day.cell(row=failures_row, column=col))
                elif cell_val == 'Time Bins (s)' or cell_val == 'Hour':
                    apply_cell_style(ws_day.cell(row=r, column=1))
                    apply_cell_style(ws_day.cell(row=r, column=2))
                    chart_data_rows.append(r) # store row number for chart creation

            # --- re-create charts for this sheet ---
            for chart_row in chart_data_rows:
                title_cell = ws_day.cell(row=chart_row, column=1).value
                chart = BarChart()
                chart.legend = None
                
                # find how many data points are in this chart table
                data_len = 0
                for i in range(1, 25): # max 24 hours or 10 bins
                    if ws_day.cell(row=chart_row + i, column=1).value is None:
                        break
                    data_len += 1
                if data_len == 0: continue

                data_ref = Reference(ws_day, min_col=2, min_row=chart_row, max_row=chart_row + data_len)
                cat_ref = Reference(ws_day, min_col=1, min_row=chart_row + 1, max_row=chart_row + data_len)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cat_ref)

                if "Distribution" in title_cell:
                    chart.title = title_cell.replace("Bins (s)", "Distribution")
                    chart.y_axis.title = 'Frequency'
                    chart.x_axis.title = 'Time Bins (s)'
                elif "Hour" in title_cell:
                    chart.title = "Hourly Package Retrieval Throughput"
                    chart.y_axis.title = 'Packages Retrieved'
                    chart.x_axis.title = 'Hour of Day'
                
                ws_day.add_chart(chart, f"E{chart_row}")

    # add final charts to master summary
    if target_sheet_name in wb.sheetnames:
        ws = wb[target_sheet_name]
        # convert data to numeric for charting
        for col_idx, col_name in enumerate(updated_summary_df.columns, 1):
            if '(s)' in col_name or 'Stowed' in col_name or 'Attempts' in col_name or 'Retrieved' in col_name or 'Errors' in col_name:
                 for row_idx, value in enumerate(updated_summary_df[col_name], 2):
                    try:
                        ws.cell(row=row_idx, column=col_idx).value = float(value)
                    except (ValueError, TypeError):
                        continue # keep 'n/a' as is

        # chart 1: average times
        chart1 = LineChart()
        chart1.title = "Average Event Times Over Time"
        chart1.y_axis.title = "Time (s)"
        chart1.x_axis.title = "Day"
        
        data1 = Reference(ws, min_col=3, max_col=5, min_row=1, max_row=ws.max_row)
        cats1 = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        
        ws.add_chart(chart1, "A" + str(ws.max_row + 2))

        # create a second chart for the errors on a secondary axis
        chart1 = LineChart()
        chart1.title = "Average Event Times Over Time"
        chart1.y_axis.title = "Time (s)"
        chart1.x_axis.title = "Day"
    
        data1 = Reference(ws, min_col=3, max_col=5, min_row=1, max_row=ws.max_row)
        cats1 = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
    
        ws.add_chart(chart1, "A" + str(ws.max_row + 2))        
        wb.move_sheet(ws, offset=-len(wb.sheetnames))
    wb.save(excel_path)

    if file_exists:
        print(f"\n✅ Master report successfully updated at: {excel_path}")
    else:
        print(f"\n✅ Master report successfully created at: {excel_path}")

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python report_updater.py <master_report.xlsx> <log_file.json> <target_sheet_name>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    log_files = [sys.argv[2]] # This script now handles one log at a time
    target_sheet = sys.argv[3]
    
    update_master_report(log_files, excel_file, target_sheet)